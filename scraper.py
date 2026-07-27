import os
import httpx
from bs4 import BeautifulSoup
import csv
import io
import json
import re
import database as db


# --- APOLLO.IO ---

async def buscar_apollo(
    empresa: str = "",
    cargo: str = "",
    ubicacion: str = "Chile",
    page: int = 1,
    per_page: int = 25,
) -> dict:
    """Busca contactos B2B via Apollo.io API."""
    api_key = os.environ.get("APOLLO_API_KEY", "")
    if not api_key:
        return {"error": "APOLLO_API_KEY no configurada. Crea cuenta gratis en apollo.io y agrega tu API key.", "people": []}

    body = {
        "api_key": api_key,
        "page": page,
        "per_page": per_page,
    }

    if empresa:
        body["q_organization_name"] = empresa
    if cargo:
        body["person_titles"] = [cargo]
    if ubicacion:
        body["person_locations"] = [ubicacion]

    headers = {
        "Content-Type": "application/json",
        "Cache-Control": "no-cache",
        "X-Api-Key": api_key,
    }
    # ponytail: Apollo deprecó api_key en body, ahora va en header
    del body["api_key"]

    async with httpx.AsyncClient(timeout=30) as client:
        try:
            # ponytail: intenta people/search (gratis), fallback a mixed_people/search (pago)
            resp = await client.post(
                "https://api.apollo.io/api/v1/people/search",
                headers=headers,
                json=body,
            )
            if resp.status_code == 403:
                resp = await client.post(
                    "https://api.apollo.io/api/v1/mixed_people/search",
                    headers=headers,
                    json=body,
                )
            if resp.status_code != 200:
                detail = resp.text[:300]
                return {"error": f"Apollo API error {resp.status_code}: {detail}", "people": []}

            data = resp.json()
        except httpx.HTTPError as e:
            return {"error": str(e), "people": []}

    people = []
    for p in data.get("people", []):
        org = p.get("organization", {}) or {}
        people.append({
            "nombre": p.get("first_name", ""),
            "apellido": p.get("last_name", ""),
            "cargo": p.get("title", ""),
            "email": p.get("email", ""),
            "telefono": (p.get("phone_numbers") or [{}])[0].get("sanitized_number", "") if p.get("phone_numbers") else "",
            "linkedin_url": p.get("linkedin_url", ""),
            "empresa": org.get("name", ""),
            "empresa_web": org.get("website_url", ""),
            "empresa_industria": org.get("industry", ""),
            "empresa_tamano": str(org.get("estimated_num_employees", "")),
            "empresa_ciudad": org.get("city", ""),
            "foto": p.get("photo_url", ""),
        })

    total = data.get("pagination", {}).get("total_entries", 0)
    db.log_scraping("apollo", f"{empresa} | {cargo} | {ubicacion}", len(people))

    return {"people": people, "total": total, "page": page}


async def importar_apollo_al_crm(contactos: list[dict]) -> dict:
    """Importa contactos de Apollo directamente al CRM."""
    empresas_creadas = 0
    contactos_creados = 0
    empresa_cache = {}

    for c in contactos:
        empresa_nombre = c.get("empresa", "")
        empresa_id = None

        if empresa_nombre:
            if empresa_nombre in empresa_cache:
                empresa_id = empresa_cache[empresa_nombre]
            else:
                empresa_id = db.crear_empresa({
                    "nombre": empresa_nombre,
                    "sitio_web": c.get("empresa_web", ""),
                    "industria": c.get("empresa_industria", ""),
                    "tamano": c.get("empresa_tamano", ""),
                    "ciudad": c.get("empresa_ciudad", ""),
                    "fuente": "apollo",
                })
                empresa_cache[empresa_nombre] = empresa_id
                empresas_creadas += 1

        if c.get("nombre"):
            db.crear_contacto({
                "empresa_id": empresa_id,
                "nombre": c["nombre"],
                "apellido": c.get("apellido", ""),
                "cargo": c.get("cargo", ""),
                "email": c.get("email", ""),
                "telefono": c.get("telefono", ""),
                "linkedin_url": c.get("linkedin_url", ""),
                "fuente": "apollo",
            })
            contactos_creados += 1

    return {"empresas_creadas": empresas_creadas, "contactos_creados": contactos_creados}


# --- HUNTER.IO ---

async def buscar_hunter(dominio: str) -> dict:
    """Busca emails de una empresa via Hunter.io (25 gratis/mes)."""
    # ponytail: limpia URL completa a solo dominio
    dominio = re.sub(r'^https?://(www\.)?', '', dominio).strip('/')
    api_key = os.environ.get("HUNTER_API_KEY", "")
    if not api_key:
        return {"error": "HUNTER_API_KEY no configurada. Crea cuenta gratis en hunter.io", "contacts": []}

    async with httpx.AsyncClient(timeout=30) as client:
        try:
            resp = await client.get(
                "https://api.hunter.io/v2/domain-search",
                params={"domain": dominio, "api_key": api_key},
            )
            if resp.status_code != 200:
                detail = resp.text[:300]
                return {"error": f"Hunter API error {resp.status_code}: {detail}", "contacts": []}

            data = resp.json().get("data", {})
        except httpx.HTTPError as e:
            return {"error": str(e), "contacts": []}

    contacts = []
    for e in data.get("emails", []):
        contacts.append({
            "nombre": e.get("first_name", ""),
            "apellido": e.get("last_name", ""),
            "cargo": e.get("position", ""),
            "email": e.get("value", ""),
            "telefono": e.get("phone_number") or "",
            "linkedin_url": e.get("linkedin", "") or "",
            "confianza": e.get("confidence", 0),
            "departamento": e.get("department", ""),
        })

    empresa = data.get("organization", dominio)
    db.log_scraping("hunter", dominio, len(contacts))

    return {
        "contacts": contacts,
        "total": data.get("total", len(contacts)),
        "empresa": empresa,
        "dominio": dominio,
        "web_url": data.get("webmail", False),
    }


async def buscar_hunter_email(nombre: str, apellido: str, dominio: str) -> dict:
    """Encuentra el email de una persona especifica via Hunter.io."""
    api_key = os.environ.get("HUNTER_API_KEY", "")
    if not api_key:
        return {"error": "HUNTER_API_KEY no configurada"}

    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.get(
            "https://api.hunter.io/v2/email-finder",
            params={"domain": dominio, "first_name": nombre, "last_name": apellido, "api_key": api_key},
        )
        if resp.status_code != 200:
            return {"error": f"Hunter error {resp.status_code}"}

        data = resp.json().get("data", {})
        return {
            "email": data.get("email", ""),
            "confianza": data.get("score", 0),
            "cargo": data.get("position", ""),
        }


async def importar_hunter_al_crm(contactos: list[dict], empresa_nombre: str, dominio: str) -> dict:
    """Importa contactos de Hunter directamente al CRM."""
    empresa_id = db.crear_empresa({
        "nombre": empresa_nombre,
        "sitio_web": f"https://{dominio}",
        "fuente": "hunter",
    })

    contactos_creados = 0
    for c in contactos:
        if c.get("nombre") or c.get("email"):
            db.crear_contacto({
                "empresa_id": empresa_id,
                "nombre": c.get("nombre", ""),
                "apellido": c.get("apellido", ""),
                "cargo": c.get("cargo", ""),
                "email": c.get("email", ""),
                "telefono": c.get("telefono", ""),
                "linkedin_url": c.get("linkedin_url", ""),
                "fuente": "hunter",
            })
            contactos_creados += 1

    return {"empresas_creadas": 1, "contactos_creados": contactos_creados}


# --- MERCADO PUBLICO ---

async def buscar_mercadopublico(keyword: str) -> dict:
    """Busca licitaciones en Mercado Publico (ChileCompra) via API oficial.

    ponytail: el sitio web de ChileCompra ahora es una SPA (Angular) y ya no
    tiene una URL de busqueda server-rendered para scrapear como fallback,
    asi que esto requiere si o si un ticket gratis de api.mercadopublico.cl.
    """
    ticket = os.environ.get("MERCADOPUBLICO_TICKET", "")
    if not ticket:
        return {
            "error": "MERCADOPUBLICO_TICKET no configurado. Solicita un ticket gratis en https://api.mercadopublico.cl/APISOCDS/OCDS/",
            "resultados": [],
        }

    async with httpx.AsyncClient(timeout=30) as client:
        try:
            resp = await client.get(
                "https://api.mercadopublico.cl/servicios/v1/publico/licitaciones.json",
                params={"keyword": keyword, "ticket": ticket},
            )
            if resp.status_code != 200:
                return {"error": f"Mercado Publico API error {resp.status_code}", "resultados": []}
            data = resp.json()
        except httpx.HTTPError as e:
            return {"error": str(e), "resultados": []}

    resultados = []
    for lic in data.get("Listado", []):
        resultados.append({
            "codigo": lic.get("CodigoExterno", ""),
            "nombre": lic.get("Nombre", ""),
            "organismo": lic.get("Comprador", {}).get("NombreOrganismo", ""),
            "estado": lic.get("CodigoEstado", ""),
            "fecha_cierre": lic.get("FechaCierre", ""),
            "monto": lic.get("MontoEstimado", 0),
        })

    db.log_scraping("mercadopublico_api", keyword, len(resultados))
    return {"resultados": resultados}


# --- SCRAPING WEB ---

async def scrape_sitio_empresa(url: str) -> dict:
    """Extrae emails y telefonos de un sitio web corporativo."""
    info = {"emails": [], "telefonos": [], "direccion": ""}
    async with httpx.AsyncClient(timeout=15, follow_redirects=True) as client:
        try:
            resp = await client.get(url, headers={"User-Agent": "Mozilla/5.0"})
        except httpx.HTTPError:
            return info

    text = resp.text
    emails = set(re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", text))
    info["emails"] = [e for e in emails if not e.endswith((".png", ".jpg", ".gif", ".css", ".js"))]

    telefonos = set(re.findall(r"(?:\+56|56)?[\s-]?(?:2|9)[\s-]?\d{4}[\s-]?\d{4}", text))
    info["telefonos"] = list(telefonos)[:10]

    soup = BeautifulSoup(text, "html.parser")
    for tag in soup.find_all(string=re.compile(r"(?:direcci[oó]n|address)", re.I)):
        parent = tag.find_parent()
        if parent:
            addr = parent.get_text(strip=True)[:200]
            if len(addr) > 10:
                info["direccion"] = addr
                break

    return info


# --- SEIA (Ministerio de Medio Ambiente) ---

# ponytail: mismo rotulo de rubro que prospector.EMPRESAS_CHILE/RUBRO_PRODUCTOS,
# para que una empresa detectada aca encaje directo en ese flujo despues.
SEIA_SECTORES = {
    "Forestal": "11",
    "Mineria": "2",
    "Energia": "7",
    "Pesca y Acuicultura": "5",
    "Construccion e Industrial": "16",
    "Transporte y Logistica": "14",
    "Alimentos y Bebidas": "16",
}


async def buscar_seia(rubro: str, limit: int = 20) -> dict:
    """Busca proyectos recientes en el SEIA (Servicio de Evaluacion Ambiental) por rubro.

    Senal proactiva: empresas que recien ingresaron un proyecto nuevo (mina, planta,
    parque de energia, etc.) antes de que aparezca cualquier licitacion en Mercado
    Publico. El sitio exige pasar primero por buscarProyectoResumen.php (fija sesion/
    referer) antes de que buscarProyectoResumenAction.php acepte la consulta real.
    """
    sector = SEIA_SECTORES.get(rubro)
    if not sector:
        return {"error": f"Rubro '{rubro}' no soportado en SEIA", "proyectos": []}

    campos = {
        "nombre": "", "titular": "", "folio": "", "selectRegion": "", "selectComuna": "",
        "tipoPresentacion": "Ambos", "projectStatus": "", "PresentacionMin": "", "PresentacionMax": "",
        "CalificaMin": "", "CalificaMax": "", "razoningreso": "", "id_tipoexpediente": "",
    }

    async with httpx.AsyncClient(timeout=30, headers={"User-Agent": "Mozilla/5.0"}) as client:
        try:
            await client.post(
                "https://seia.sea.gob.cl/busqueda/buscarProyectoResumen.php",
                data={**campos, "sectores_economicos[]": sector},
            )
            resp = await client.post(
                "https://seia.sea.gob.cl/busqueda/buscarProyectoResumenAction.php",
                data={**campos, "sectores_economicos": sector, "offset": 1, "limit": limit,
                      "orderColumn": "FECHA_PRESENTACION", "orderDir": "desc"},
                headers={"Referer": "https://seia.sea.gob.cl/busqueda/buscarProyectoResumen.php"},
            )
            # ponytail: el SEIA declara charset=ISO-8859-1 pero httpx .json() decodifica
            # los bytes crudos asumiendo UTF-8 e ignora el charset del header -> revienta
            # con tildes/enie. Se decodifica explicito antes de parsear.
            data = json.loads(resp.content.decode("iso-8859-1"))
        except httpx.HTTPError as e:
            return {"error": str(e), "proyectos": []}
        except Exception:
            return {"error": "Respuesta inesperada del SEIA, intenta de nuevo", "proyectos": []}

    if not data.get("status"):
        return {"error": "El SEIA no acepto la consulta, intenta de nuevo", "proyectos": []}

    proyectos = [{
        "titular": p.get("TITULAR", ""),
        "proyecto": p.get("EXPEDIENTE_NOMBRE", ""),
        "region": p.get("REGION_NOMBRE", ""),
        "estado": p.get("ESTADO_PROYECTO", ""),
        "inversion_mm": p.get("INVERSION_MM_FORMAT", ""),
        "fecha": p.get("FECHA_PRESENTACION_FORMAT", ""),
        "tipologia": p.get("DESCRIPCION_TIPOLOGIA", ""),
        "url_ficha": p.get("EXPEDIENTE_URL_FICHA", ""),
    } for p in data.get("data", [])]

    db.log_scraping("seia", rubro, len(proyectos))
    return {"proyectos": proyectos}


async def obtener_contacto_seia(url_ficha: str) -> dict:
    """Extrae la seccion 'Informacion de Contacto' de la ficha de un proyecto SEIA:
    Titular (la empresa), Representante Legal, Empresa Consultora Ambiental y
    Consultor/a -- cada uno con nombre/domicilio/ciudad/telefono/email cuando
    esten disponibles. Esta es la fuente real de correos y nombres por proyecto
    que pidio Diego, no hace falta parsear los PDF de los expedientes.
    """
    async with httpx.AsyncClient(timeout=20, follow_redirects=True, headers={"User-Agent": "Mozilla/5.0"}) as client:
        try:
            resp = await client.get(url_ficha)
        except httpx.HTTPError as e:
            return {"error": str(e)}

    try:
        soup = BeautifulSoup(resp.content.decode("iso-8859-1"), "html.parser")
    except Exception:
        return {"error": "No se pudo leer la ficha"}

    secciones = {}
    for item in soup.select(".accordion-item"):
        header = item.select_one(".accordion-button")
        if not header:
            continue
        tipo = header.get_text(strip=True)
        datos = {}
        for row in item.select(".row.sg-row-file-description"):
            cols = row.select(":scope > div")
            if len(cols) >= 2:
                label = cols[0].get_text(strip=True)
                valor = cols[1].get_text(strip=True)
                if valor:
                    datos[label] = valor
        if datos:
            secciones[tipo] = datos

    return {"contacto": secciones}


# --- IMPORTACION CSV/EXCEL ---

def importar_csv(contenido: bytes, fuente: str = "csv") -> dict:
    text = contenido.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    field_map = {}
    if reader.fieldnames:
        for f in reader.fieldnames:
            fl = f.lower().strip()
            if "empresa" in fl or "company" in fl or "razon" in fl:
                field_map["empresa"] = f
            elif "nombre" in fl or "first" in fl or "name" in fl:
                field_map["nombre"] = f
            elif "apellido" in fl or "last" in fl:
                field_map["apellido"] = f
            elif "cargo" in fl or "position" in fl or "title" in fl:
                field_map["cargo"] = f
            elif "email" in fl or "correo" in fl or "mail" in fl:
                field_map["email"] = f
            elif "tel" in fl or "phone" in fl or "fono" in fl:
                field_map["telefono"] = f
            elif "rut" in fl:
                field_map["rut"] = f
            elif "industria" in fl or "industry" in fl or "sector" in fl or "rubro" in fl:
                field_map["industria"] = f
            elif "linkedin" in fl:
                field_map["linkedin"] = f
            elif "ciudad" in fl or "city" in fl:
                field_map["ciudad"] = f

    empresas_creadas = 0
    contactos_creados = 0
    empresa_cache = {}

    for row in reader:
        empresa_nombre = row.get(field_map.get("empresa", ""), "").strip()
        empresa_id = None
        if empresa_nombre:
            if empresa_nombre in empresa_cache:
                empresa_id = empresa_cache[empresa_nombre]
            else:
                empresa_id = db.crear_empresa({
                    "nombre": empresa_nombre,
                    "rut": row.get(field_map.get("rut", ""), ""),
                    "industria": row.get(field_map.get("industria", ""), ""),
                    "ciudad": row.get(field_map.get("ciudad", ""), ""),
                    "fuente": fuente,
                })
                empresa_cache[empresa_nombre] = empresa_id
                empresas_creadas += 1

        nombre = row.get(field_map.get("nombre", ""), "").strip()
        if nombre:
            db.crear_contacto({
                "empresa_id": empresa_id,
                "nombre": nombre,
                "apellido": row.get(field_map.get("apellido", ""), ""),
                "cargo": row.get(field_map.get("cargo", ""), ""),
                "email": row.get(field_map.get("email", ""), ""),
                "telefono": row.get(field_map.get("telefono", ""), ""),
                "linkedin_url": row.get(field_map.get("linkedin", ""), ""),
                "fuente": fuente,
            })
            contactos_creados += 1

    db.log_scraping(fuente, "importacion_csv", empresas_creadas + contactos_creados)
    return {"empresas_creadas": empresas_creadas, "contactos_creados": contactos_creados, "columnas_detectadas": field_map}


def importar_excel(contenido: bytes, fuente: str = "excel") -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(contenido), read_only=True)
    ws = wb.active
    output = io.StringIO()
    writer = csv.writer(output)
    for row in ws.iter_rows(values_only=True):
        writer.writerow([str(c) if c is not None else "" for c in row])
    return importar_csv(output.getvalue().encode("utf-8"), fuente)
